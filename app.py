from flask import Flask, render_template
from flask import request, redirect, url_for, jsonify
from backend.db import get_connection
from backend.routes.tool_routes import tool_routes
from backend.routes.maintenance_routes import maintenance_routes
from backend.routes.receive_routes import receive_routes
from backend.routes.received_routes import received_routes

app = Flask(__name__, template_folder="templates", static_folder="static")

# Register blueprints for API logic (for AJAX or minimal JS)
app.register_blueprint(tool_routes, url_prefix="/api/tools")
app.register_blueprint(maintenance_routes, url_prefix="/api/maintenance")
app.register_blueprint(receive_routes, url_prefix="/api")
app.register_blueprint(received_routes, url_prefix="/api/received")

@app.route("/")
def dashboard():
    return render_template("dashboard.html")

# Register Tool
@app.route("/register", methods=["GET", "POST"])
def register_tool():
    message = None
    if request.method == "POST":
        rfid_uid = request.form.get("rfid_uid", "").strip()
        tool_name = request.form.get("tool_name", "").strip()
        model_no = request.form.get("model_no", "").strip()
        if not rfid_uid or not tool_name:
            message = "RFID UID and Tool Name are required"
        else:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT rfid_uid FROM tools WHERE rfid_uid = %s", (rfid_uid,))
            rfid_exists = cursor.fetchall()
            if len(rfid_exists) > 0:
                message = "RFID already registered"
            else:
                final_name = tool_name
                counter = 1
                while True:
                    cursor.execute("SELECT tool_name FROM tools WHERE tool_name = %s", (final_name,))
                    name_check = cursor.fetchall()
                    if len(name_check) == 0:
                        break
                    final_name = f"{tool_name}{counter}"
                    counter += 1
                cursor.execute(
                    "INSERT INTO tools (rfid_uid, tool_name, model_no) VALUES (%s, %s, %s)",
                    (rfid_uid, final_name, model_no if model_no else None)
                )
                conn.commit()
                message = "Tool registered successfully"
            cursor.close()
            conn.close()
    return render_template("register_tool.html", message=message)

# Scan Maintenance
@app.route("/scan", methods=["GET", "POST"])
def scan_maintenance():
    message = None
    if request.method == "POST":
        rfid_uid = request.form.get("rfid_uid", "").strip()
        if not rfid_uid:
            message = "RFID UID is required"
        else:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT rfid_uid, tool_name, model_no FROM tools WHERE rfid_uid = %s", (rfid_uid,))
            tool_rows = cursor.fetchall()
            if len(tool_rows) == 0:
                message = "Tool not registered"
            else:
                cursor.execute("SELECT rfid_uid FROM maintenance_tools WHERE rfid_uid = %s", (rfid_uid,))
                maint_rows = cursor.fetchall()
                if len(maint_rows) > 0:
                    message = "Tool already under maintenance"
                else:
                    cursor.execute("SELECT id FROM received_tools WHERE rfid_uid = %s AND received_date IS NULL", (rfid_uid,))
                    history = cursor.fetchall()
                    if len(history) == 0:
                        cursor.execute(
                            "INSERT INTO received_tools (rfid_uid, tool_name, model_no) VALUES (%s, %s, %s)",
                            (tool_rows[0]['rfid_uid'], tool_rows[0]['tool_name'], tool_rows[0]['model_no'])
                        )
                    cursor.execute(
                        "INSERT INTO maintenance_tools (rfid_uid, tool_name, model_no) VALUES (%s, %s, %s)",
                        (tool_rows[0]['rfid_uid'], tool_rows[0]['tool_name'], tool_rows[0]['model_no'])
                    )
                    conn.commit()
                    message = "Tool sent to maintenance successfully"
            cursor.close()
            conn.close()
    return render_template("scan_maintenance.html", message=message)

# Receive Maintenance
@app.route("/receive", methods=["GET", "POST"])
def receive_maintenance():
    message = None
    if request.method == "POST":
        rfid_uid = request.form.get("rfid_uid", "").strip()
        if not rfid_uid:
            message = "RFID is required"
        else:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT rfid_uid FROM maintenance_tools WHERE rfid_uid = %s", (rfid_uid,))
            active = cursor.fetchall()
            if len(active) == 0:
                message = "Tool is not under maintenance"
            else:
                cursor.execute(
                    "UPDATE received_tools SET received_date = NOW() WHERE rfid_uid = %s AND received_date IS NULL",
                    (rfid_uid,)
                )
                cursor.execute("DELETE FROM maintenance_tools WHERE rfid_uid = %s", (rfid_uid,))
                conn.commit()
                message = "Tool received successfully"
            cursor.close()
            conn.close()
    return render_template("receive_maintenance.html", message=message)


@app.route("/registered_tools", methods=["GET"])
def registered_tools():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT rfid_uid, tool_name, model_no FROM tools ORDER BY tool_name ASC")
    tools = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template("registered_tools.html", tools=tools)

# Maintenance History with search, date filter, and PDF download
from flask import send_file, request
import io
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@app.route("/history")
def received_tools():
    search = request.args.get('search', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    sort_by = request.args.get('sort_by', '').strip()
    query = "SELECT rfid_uid, tool_name, model_no, sent_date, received_date FROM received_tools WHERE 1=1"
    params = []
    if search:
        query += " AND tool_name LIKE %s"
        params.append(f"%{search}%")
    if start_date:
        query += " AND DATE(sent_date) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND DATE(sent_date) <= %s"
        params.append(end_date)
    # Apply sort
    if sort_by == 'sent_date_asc':
        query += " ORDER BY sent_date ASC"
    elif sort_by == 'received_date_desc':
        query += " ORDER BY received_date DESC"
    elif sort_by == 'received_date_asc':
        query += " ORDER BY received_date ASC"
    elif sort_by == 'tool_name':
        query += " ORDER BY tool_name ASC"
    else:
        query += " ORDER BY sent_date DESC"
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    tools = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template("received_tools.html", tools=tools)

# PDF download route
@app.route("/history/pdf")
def download_pdf():
    search = request.args.get('search', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    sort_by = request.args.get('sort_by', '').strip()
    query = "SELECT rfid_uid, tool_name, model_no, sent_date, received_date FROM received_tools WHERE 1=1"
    params = []
    if search:
        query += " AND tool_name LIKE %s"
        params.append(f"%{search}%")
    if start_date:
        query += " AND DATE(sent_date) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND DATE(sent_date) <= %s"
        params.append(end_date)
    # Apply sort
    if sort_by == 'sent_date_asc':
        query += " ORDER BY sent_date ASC"
    elif sort_by == 'received_date_desc':
        query += " ORDER BY received_date DESC"
    elif sort_by == 'received_date_asc':
        query += " ORDER BY received_date ASC"
    elif sort_by == 'tool_name':
        query += " ORDER BY tool_name ASC"
    else:
        query += " ORDER BY sent_date DESC"
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    tools = cursor.fetchall()
    cursor.close()
    conn.close()
    # Generate PDF with Table format
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0a2342'), spaceAfter=12, alignment=1)
    story.append(Paragraph('Maintenance History Report', title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Prepare table data
    table_data = [['RFID UID', 'Tool Name', 'Model No', 'Sent Date', 'Received Date']]
    for tool in tools:
        table_data.append([
            str(tool['rfid_uid']),
            str(tool['tool_name']) if tool['tool_name'] else 'N/A',
            str(tool['model_no']) if tool['model_no'] else 'N/A',
            str(tool['sent_date']) if tool['sent_date'] else 'N/A',
            str(tool['received_date']) if tool['received_date'] else 'N/A'
        ])
    
    # Create table with styling
    table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 1.3*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafd')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="maintenance_history.pdf", mimetype="application/pdf")

# Excel download route
@app.route("/download_excel")
def download_excel():
    search = request.args.get('search', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    sort_by = request.args.get('sort_by', '').strip()
    query = "SELECT rfid_uid, tool_name, model_no, sent_date, received_date FROM received_tools WHERE 1=1"
    params = []
    if search:
        query += " AND tool_name LIKE %s"
        params.append(f"%{search}%")
    if start_date:
        query += " AND DATE(sent_date) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND DATE(sent_date) <= %s"
        params.append(end_date)
    # Apply sort
    if sort_by == 'sent_date_asc':
        query += " ORDER BY sent_date ASC"
    elif sort_by == 'received_date_desc':
        query += " ORDER BY received_date DESC"
    elif sort_by == 'received_date_asc':
        query += " ORDER BY received_date ASC"
    elif sort_by == 'tool_name':
        query += " ORDER BY tool_name ASC"
    else:
        query += " ORDER BY sent_date DESC"
    
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    tools = cursor.fetchall()
    cursor.close()
    conn.close()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance History"
    
    # Add headers
    headers = ['RFID UID', 'Tool Name', 'Model No', 'Sent Date', 'Received Date']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for tool in tools:
        ws.append([
            tool['rfid_uid'],
            tool['tool_name'],
            tool['model_no'],
            str(tool['sent_date']),
            str(tool['received_date']) if tool['received_date'] else 'Pending'
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Center align all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name="maintenance_history.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Safe bulk delete for Registered Tools
from flask import jsonify

@app.route("/delete_selected_tools", methods=["POST"])
def delete_selected_tools():
    data = request.get_json()
    rfids = data.get("rfids", [])
    if not rfids or not isinstance(rfids, list):
        return jsonify({"success": False, "message": "No tools selected."}), 400
    conn = get_connection()
    cursor = conn.cursor()
    deleted = 0
    try:
        for rfid in rfids:
            # Delete ONLY from tools table (registered tools list)
            # This removes it from the registered tools but keeps all history intact
            cursor.execute("DELETE FROM tools WHERE rfid_uid = %s", (rfid,))
            if cursor.rowcount:
                deleted += 1
            # Note: maintenance_tools and received_tools records are NOT deleted
            # This preserves the complete maintenance history
        
        conn.commit()
        return jsonify({
            "success": True,
            "deleted": deleted,
            "message": f"Successfully deleted {deleted} tool(s). All maintenance history preserved."
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route("/delete_history_records", methods=["POST"])
def delete_history_records():
    data = request.get_json()
    records = data.get("records", [])  # List of {rfid_uid, sent_date} objects
    if not records or not isinstance(records, list):
        return jsonify({"success": False, "message": "No records selected."}), 400
    conn = get_connection()
    cursor = conn.cursor()
    deleted = 0
    try:
        for record in records:
            rfid_uid = record.get("rfid_uid")
            sent_date = record.get("sent_date")
            
            if not rfid_uid or not sent_date:
                continue
            
            # Delete only the specific received_tools record by rfid_uid AND sent_date combination
            cursor.execute(
                "DELETE FROM received_tools WHERE rfid_uid = %s AND sent_date = %s",
                (rfid_uid, sent_date)
            )
            deleted += cursor.rowcount
        conn.commit()
        return jsonify({"success": True, "deleted": deleted})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    app.run(debug=True)
